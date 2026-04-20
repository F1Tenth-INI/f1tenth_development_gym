#!/usr/bin/env python3
"""
Create a ranking table visualization from learning metrics leaderboards.
Generates both Excel (with color coding) and PNG versions.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.colors import LinearSegmentedColormap, Normalize
import seaborn as sns

# Configuration
DATA_DIR = Path("batch_learning_metrics_summary/RCA2-Fixed")
CSV_FILE = DATA_DIR / "learning_metric_leaderboards_RCA2-Fixed.csv"
OUTPUT_EXCEL = DATA_DIR / "ranking_table_RCA2-Fixed.xlsx"
OUTPUT_PNG = DATA_DIR / "ranking_table_RCA2-Fixed.png"

def load_and_process_data(csv_path):
    """Load CSV and create ranking table."""
    df = pd.read_csv(csv_path)
    
    # Get unique metrics and models
    metrics = df['metric_id'].unique()
    models = df['name'].unique()
    
    print(f"Found {len(models)} models and {len(metrics)} metrics")
    print(f"Metrics: {sorted(metrics)}")
    
    # Create pivot table: rows=models, columns=metrics, values=rank
    ranking_table = df.pivot_table(
        index='name', 
        columns='metric_id', 
        values='rank',
        aggfunc='first'
    )
    
    # Reorder columns logically
    ranking_table = ranking_table.reindex(sorted(ranking_table.columns), axis=1)
    
    # Calculate overall ranking (lower sum of ranks = better)
    # Handle NaN values by treating them as the worst rank + 1
    overall_points = ranking_table.fillna(len(models) + 1).sum(axis=1)
    ranking_table['Overall_Rank'] = overall_points.rank(method='min').astype(int)
    
    # Sort by overall rank
    ranking_table = ranking_table.sort_values('Overall_Rank')
    
    return ranking_table, metrics

def create_excel_with_colors(ranking_table, output_path):
    """Create Excel file with conditional formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rankings"
    
    # Write headers
    ws['A1'] = "Model"
    for i, col in enumerate(ranking_table.columns, 1):
        col_letter = get_column_letter(i + 1)
        ws[f'{col_letter}1'] = col
    
    # Write data
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Apply header styling
    for cell in ws[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    # Write rows with data
    for row_idx, (model_name, row_data) in enumerate(ranking_table.iterrows(), 2):
        # Model name
        ws[f'A{row_idx}'] = model_name
        ws[f'A{row_idx}'].border = thin_border
        ws[f'A{row_idx}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Metric ranks
        for col_idx, value in enumerate(row_data, 2):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{row_idx}']
            
            # Handle NaN (not reached metric)
            if pd.isna(value):
                cell.value = "-"
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.font = Font(color="808080")
            else:
                cell.value = int(value)
                # Color code: green for rank 1, red for last rank (lower is better)
                num_models = len(ranking_table)
                # Invert so rank 1 gets highest score (closest to 1.0)
                norm_rank = (num_models - value) / (num_models - 1) if num_models > 1 else 1.0
                
                # Red to Yellow to Green gradient
                if norm_rank < 0.5:
                    # Red to Yellow
                    r = 255
                    g = int(255 * (norm_rank * 2))
                    b = 0
                else:
                    # Yellow to Green
                    r = int(255 * (2 - norm_rank * 2))
                    g = 255
                    b = 0
                
                color_hex = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                # Use black text for better readability
                cell.font = Font(bold=True, color="000000")
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 60
    for i in range(2, len(ranking_table.columns) + 2):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 15
    
    # Freeze header and first column
    ws.freeze_panes = 'B2'
    
    wb.save(output_path)
    print(f"✓ Excel file saved: {output_path}")

def create_png_visualization(ranking_table, output_path):
    """Create PNG heatmap visualization."""
    # Prepare data for visualization
    data_for_viz = ranking_table.copy()
    
    # Create color array (NaN values will be gray)
    color_array = np.full_like(data_for_viz.values, np.nan, dtype=float)
    for i in range(data_for_viz.shape[0]):
        for j in range(data_for_viz.shape[1]):
            if not pd.isna(data_for_viz.iloc[i, j]):
                color_array[i, j] = data_for_viz.iloc[i, j]
    
    # Create figure
    fig, ax = plt.subplots(figsize=(40, 20))
    
    num_models = len(data_for_viz)
    num_metrics = len(data_for_viz.columns)
    
    # Scale factor to spread out columns horizontally
    col_scale = 3  # Multiply column positions and widths by this factor
    
    # Create heatmap manually to handle NaN with gray
    max_rank = num_models
    
    for i in range(num_models):
        for j in range(num_metrics):
            value = data_for_viz.iloc[i, j]
            
            if pd.isna(value):
                # Gray for missing values
                color = '#D3D3D3'
                text = '-'
            else:
                # Use same RGB calculation as Excel (lower rank = greener)
                norm_rank = (max_rank - value) / (max_rank - 1) if max_rank > 1 else 1.0
                
                # Red to Yellow to Green gradient
                if norm_rank < 0.5:
                    # Red to Yellow
                    r = 255
                    g = int(255 * (norm_rank * 2))
                    b = 0
                else:
                    # Yellow to Green
                    r = int(255 * (2 - norm_rank * 2))
                    g = 255
                    b = 0
                
                color = f'#{r:02X}{g:02X}{b:02X}'
                text = str(int(value))
            
            # Scale the rectangle positions and width
            rect = mpatches.Rectangle((j * col_scale, num_models - i - 1), col_scale * 0.9, 1, 
                                     linewidth=0.5, edgecolor='white',
                                     facecolor=color)
            ax.add_patch(rect)
            
            # Add text
            ax.text(j * col_scale + col_scale * 0.45, num_models - i - 1 + 0.5, text,
                   ha='center', va='center', fontsize=7, fontweight='bold')
    
    # Set axis properties
    ax.set_xlim(0, num_metrics * col_scale)
    ax.set_ylim(0, num_models)
    ax.set_aspect('equal')
    
    # Labels
    ax.set_xticks([j * col_scale + col_scale * 0.45 for j in range(num_metrics)])
    ax.set_xticklabels(data_for_viz.columns, rotation=45, ha='right', fontsize=9)
    
    ax.set_yticks(np.arange(num_models) + 0.5)
    model_labels = [name[:40] + '...' if len(name) > 40 else name 
                    for name in data_for_viz.index]
    ax.set_yticklabels(model_labels, fontsize=7)
    
    # Title
    ax.set_title('Model Ranking Heatmap\n(Green=Best Rank, Red=Worst Rank, Gray=Not Reached)', 
                fontsize=14, fontweight='bold', pad=20)
    
    # Add colorbar legend
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='#FF0000', edgecolor='black', label='Worst Rank'),
        Patch(facecolor='#FFFF00', edgecolor='black', label='Medium Rank'),
        Patch(facecolor='#00FF00', edgecolor='black', label='Best Rank (1)'),
        Patch(facecolor='#D3D3D3', edgecolor='black', label='Not Reached')
    ]
    ax.legend(handles=legend_elements, loc='upper left', fontsize=9)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    print(f"✓ PNG file saved: {output_path}")
    plt.close()

def main():
    """Main execution."""
    print("Loading data...")
    ranking_table, metrics = load_and_process_data(CSV_FILE)
    
    print(f"\nRanking table shape: {ranking_table.shape}")
    print(f"\nTop 5 models:")
    print(ranking_table.head(5))
    
    print("\nCreating Excel file...")
    create_excel_with_colors(ranking_table, OUTPUT_EXCEL)
    
    print("Creating PNG visualization...")
    create_png_visualization(ranking_table, OUTPUT_PNG)
    
    print("\n✓ Done! Files created:")
    print(f"  - {OUTPUT_EXCEL}")
    print(f"  - {OUTPUT_PNG}")

if __name__ == "__main__":
    main()
